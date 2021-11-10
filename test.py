import sqlite3
import re
from openpyxl import load_workbook
import dictionary_template

"""-------------------diverse function-----------------"""


def thickness_cut(thickness):
    return thickness[10:-10]


def melting_number_cut(melting_number):
    if melting_number is not None:
        return melting_number.split()[0]
    else:
        melting_number = 'Error Cell None'
        return melting_number


def batch_number_cut(batch_number):
    if batch_number is not None:
        if len(batch_number.split()) > 1:
            # return batch_number.split()[1]
            return re.sub(r"[(,)]", "", batch_number.split()[1])
        else:
            return ' '
    else:
        batch_number = 'Error Cell None'
        return batch_number


"""----------------end diverse function-----------------"""


"""----------------working with a table-----------------"""
START_ROW = 5  # constant start read xls file
end_row = 5
my_workbook = load_workbook('C:/Копия журнал верификации лист 09Г2С - копия.xlsx')
my_worksheet = my_workbook.active

while my_worksheet.cell(end_row, 4).value:  # find the last non-empty cell
    end_row += 1

print(end_row)
"""------------end working with a table-----------------"""

my_db = []
count = 0
for i in range(START_ROW, end_row):
    row = ()
    row += (count, )
    for key in dictionary_template.dictionary_template.keys():
        row = row + (my_worksheet.cell(i, dictionary_template.dictionary_template[key]).value, )
        #row[key] = my_worksheet.cell(i, dictionary_template.dictionary_template[key]).value
    count += 1
    my_db.append(row)

conn = sqlite3.connect('cert_db.db')
cur = conn.cursor()

cur.execute("""CREATE TABLE IF NOT EXISTS cert_db(
   cert_id INTEGER PRIMARY KEY,
   thickness TEXT,
   metal_grade TEXT,
   gost TEXT,
   melting_number TEXT,
   batch_number TEXT,
   certificate_number TEXT,
   date_of_certificate TEXT,
   limit_fluidity TEXT,
   temporary_resistance TEXT,
   relative_extension TEXT,
   relative_narrowing TEXT,
   impact_strength_before_aging TEXT,
   impact_strength_after_aging TEXT,
   sample_type TEXT,
   impact_strength_below_zero TEXT,
   temperature_below_zero TEXT,
   sample_type_below_zero TEXT,
   additional_data TEXT);
""")

conn.commit()

print(len(my_db[0]))
row_element_count = ''
for i in range(0, len(my_db[0])):
    if i != len(my_db[0])-1:
        row_element_count += '?, '
    else:
        row_element_count += '?'
print(row_element_count)
cur.executemany(f"INSERT INTO cert_db VALUES({row_element_count});", my_db)

cur.execute("SELECT thickness, COUNT(melting_number) FROM cert_db GROUP BY thickness ORDER BY 2 DESC;")
one_result = cur.fetchall()
for row_p in one_result:
    print(row_p)
