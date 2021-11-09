from tkinter import *
import string

from openpyxl import load_workbook

root = Tk()
root.title('openXL')


my_workbook = load_workbook('C:/Копия журнал верификации лист 09Г2С - копия.xlsx')
my_worksheet = my_workbook.active
print(my_worksheet.max_row)

thickness = 100
count = 0
column_template = [3, 5, 6, 7, 8, 9]

for my_row in range(5, my_worksheet.max_row):
    if not my_worksheet['A' + str(my_row)].value:
        break
    me_row = my_worksheet[my_row]
    for value in me_row:
        print(value.value, end='][')
    print()

# for x in range(6, my_worksheet.max_row):
#     count += 1
#     thickness_search = my_worksheet['D' + str(x)].value
#     if not thickness_search:
#         break
#     thickness_search = thickness_search[8:-10].strip()
#     if thickness_search == str(thickness):
#         print('[', count,'] ', my_worksheet['M' + str(x)].value)
#     # print(thickness_search)
#     # print(my_worksheet['A' + str(x)].value)


h = 'Лист г/к  8*1500*6000'
y = h.split('Лист г/к')
h = h[8:-10].strip()


print(y)


print(h)
# print(my_worksheet.cell(row = 6, column = column_s).value, end=' ---  ')
# for x in range(6, my_worksheet.max_row):
#     if not my_worksheet['M' + str(x)].value:
#         break
#     print(my_worksheet['M' + str(x)].value.split())


root.mainloop()

"""
{
  "melting_number": "?",
  "certificate_number": "?",
  "date_of_certificate": "?",
  "metal_grade": "?",
  "thickness": "?"
}
"""